#!/usr/bin/env python3
# Wiltek Portal — V1 第 2/2.1/3 刀 today + customer-insights data prep
# Reads:
#   /Users/jymchee/Desktop/Claude use/202604 CUSTOMER BUY.V3.xlsx
#   /Users/jymchee/Desktop/Claude use/202604 - SALES VS STOCK VS PO VS GRN.V2.xlsx
# Emits:
#   /Users/jymchee/wiltek-repo/assets/today-data.js      (today + churn + po_exceptions)
#   /Users/jymchee/wiltek-repo/assets/customers-data.js  (V1 第 3 刀 RFM + cross-table + top100)
import json, os
from collections import defaultdict, Counter
from datetime import datetime
import openpyxl

CUST_XLSX = "/Users/jymchee/Desktop/Claude use/202604 CUSTOMER BUY.V3.xlsx"
SVS_XLSX  = "/Users/jymchee/Desktop/Claude use/202604 - SALES VS STOCK VS PO VS GRN.V2.xlsx"
OUT_TODAY = "/Users/jymchee/wiltek-repo/assets/today-data.js"
OUT_CUST  = "/Users/jymchee/wiltek-repo/assets/customers-data.js"

NOW_YM = (2026, 3)
NOW_DT = datetime(2026, 3, 31)
LTM_CUTOFF = datetime(2025, 4, 1)   # last twelve months from snapshot
M6_CUTOFF  = datetime(2025, 10, 1)  # last 6 months
M3_CUTOFF  = datetime(2026, 1, 1)   # last 3 months
M1_CUTOFF  = datetime(2026, 3, 1)   # last 1 month

ACTIVE_BRANCHES = {"W01", "W02", "W03", "W05", "W07"}

# V1 第8刀 (2026-05-06) — Walk-in by race is now read LIVE from 5 W0X
# Customer Floatation Sheets via /api/floatation. The hardcoded literal that
# used to live below has been removed; the build no longer emits a `race` key
# into customers-data.js. Front-end fetches the live shape on page load
# (and on every Refresh click) and mutates window.WP_TODAY.race in place.
#
# Kept the variable name + signature as a no-op stub so anyone reading old
# diffs sees the deletion explicitly rather than a silent vanish.
RACE_DATA = None  # live now — see api/floatation.js
_DELETED_RACE_DATA = {
    "months": ["2026-01", "2026-02", "2026-03"],
    "races": [
        {"key": "chinese",  "label_en": "Chinese",  "label_zh": "华族",
         "walkin": [683, 468, 565], "purchase": [531, 367, 407],
         "amount": [179646, 130036.38, 136072.60],
         "basket": [338.32, 354.32, 334.33], "cr": [0.7775, 0.7842, 0.7204]},
        {"key": "malay",    "label_en": "Malay",    "label_zh": "马来族",
         "walkin": [896, 825, 764], "purchase": [692, 652, 589],
         "amount": [300092.10, 267007.32, 224020],
         "basket": [433.66, 409.52, 380.34], "cr": [0.7723, 0.7903, 0.7709]},
        {"key": "indian",   "label_en": "Indian",   "label_zh": "印度族",
         "walkin": [56, 48, 61], "purchase": [48, 41, 52],
         "amount": [18428, 12941.60, 11448],
         "basket": [383.92, 315.65, 220.15], "cr": [0.8571, 0.8542, 0.8525]},
        {"key": "others",   "label_en": "Others",   "label_zh": "其他",
         "walkin": [23, 24, 27], "purchase": [20, 21, 22],
         "amount": [6063.50, 3573, 7388.05],
         "basket": [303.18, 170.14, 335.82], "cr": [0.8696, 0.875, 0.8148]},
    ],
    "totals": {
        "walkin":   [1658, 1365, 1417],
        "purchase": [1291, 1081, 1070],
        "amount":   [504229.60, 413558.30, 378928.65],
        "basket":   [390.57, 382.57, 354.14],
        "cr":       [0.7786, 0.7919, 0.7551],
    },
    "by_branch": {
        "W01": {"walkin": 408, "purchase": 306, "amount": 98926,  "basket": 323.29, "cr": 0.75},
        "W02": {"walkin": 551, "purchase": 391, "amount": 151629, "basket": 387.80, "cr": 0.7096},
        "W03": {"walkin": 427, "purchase": 304, "amount": 142021, "basket": 467.17, "cr": 0.7119},
        "W05": {"walkin": 296, "purchase": 265, "amount": 122197, "basket": 461.12, "cr": 0.8953},
        "W07": {"walkin": 614, "purchase": 436, "amount": 161534, "basket": 370.49, "cr": 0.71},
    },
    "note_en": "Manual count of foot traffic by race, Jan–Mar 2026. POS doesn't track race — branches log this separately. WCO/W11 excluded.",
    "note_zh": "各店人手记录的进店种族数据,2026年1-3月。POS 不记种族,各店单独统计。WCO/W11 不计入。",
}

# Cust-type code → display label (raw codes in xlsx are 1-letter)
CT_LABEL = {
    "N":   "Walk-in",
    "C":   "Contractor",
    "D":   "Interior Designer",
}
def ct_norm(raw):
    if raw is None: return "Other"
    s = str(raw).strip().upper()
    if s in CT_LABEL: return CT_LABEL[s]
    return "Other"

def shift(ym, n):
    y, m = ym
    total = y * 12 + (m - 1) - n
    return (total // 12, total % 12 + 1)

def ym_str(ym):
    return f"{ym[0]:04d}-{ym[1]:02d}"

# ── Customer master pass (collect everything we need across all 3 cuts) ──
print("Reading customer xlsx (single pass) …")
wb = openpyxl.load_workbook(CUST_XLSX, read_only=True, data_only=True)
ws = wb["Sheet27"]
mem = defaultdict(lambda: {
    "last": None, "first": None, "amt": 0.0,
    "visits": set(), "name": "", "branches": defaultdict(float), "loy": "",
    "enrol": None, "cust_type": "",
    "ltm_amt": 0.0, "ltm_visits": set(),
    "m6_amt": 0.0,  "m6_visits":  set(),
    "m3_amt": 0.0,  "m3_visits":  set(),
    "m1_amt": 0.0,  "m1_visits":  set(),
})
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or r[0] is None: continue
    month, bill, code, br, name, mc, qty, amt, ct, enrol, loy, mg, sg = r
    if not mc: continue
    if not isinstance(month, datetime): continue
    d = mem[mc]
    if d["last"] is None or month > d["last"]: d["last"] = month
    if d["first"] is None or month < d["first"]: d["first"] = month
    amt_f = float(amt or 0)
    d["amt"] += amt_f
    d["visits"].add(bill)
    if not d["name"]: d["name"] = name or ""
    if br: d["branches"][br] += amt_f
    if loy: d["loy"] = loy
    # Date Enrol — keep earliest non-null sane value
    if isinstance(enrol, datetime) and (d["enrol"] is None or enrol < d["enrol"]):
        # filter out garbage like 2099 placeholders
        if 1990 <= enrol.year <= 2026:
            d["enrol"] = enrol
    if ct and not d["cust_type"]: d["cust_type"] = ct_norm(ct)
    # Window aggregates (LTM / 6m / 3m / 1m from snapshot)
    if month >= LTM_CUTOFF:
        d["ltm_amt"] += amt_f
        d["ltm_visits"].add(bill)
    if month >= M6_CUTOFF:
        d["m6_amt"] += amt_f
        d["m6_visits"].add(bill)
    if month >= M3_CUTOFF:
        d["m3_amt"] += amt_f
        d["m3_visits"].add(bill)
    if month >= M1_CUTOFF:
        d["m1_amt"] += amt_f
        d["m1_visits"].add(bill)
wb.close()
print(f"  Total members: {len(mem):,}")

# ── Customer churn (D1) ──────────────────────────────────────────────
CHURN_CUTOFF = shift(NOW_YM, 5)
churned = []
for mc, d in mem.items():
    last_ym = (d["last"].year, d["last"].month)
    if last_ym >= CHURN_CUTOFF: continue
    if d["amt"] < 500: continue
    if len(d["visits"]) < 2: continue
    primary_branch = max(d["branches"].items(), key=lambda x: x[1])[0]
    months_ago = (NOW_YM[0] - last_ym[0]) * 12 + (NOW_YM[1] - last_ym[1])
    churned.append({
        "mc": str(mc),
        "name": (d["name"][:40] if d["name"] else str(mc)),
        "last": ym_str(last_ym),
        "months_ago": months_ago,
        "amount": round(d["amt"], 0),
        "visits": len(d["visits"]),
        "loyalty": d["loy"],
        "branch": primary_branch,
        "cust_type": d["cust_type"] or "Other",
    })
churned.sort(key=lambda x: -x["amount"])
TIER_HIGH_THRESHOLD = 1000
high_value_churn = [c for c in churned if c["amount"] >= TIER_HIGH_THRESHOLD]
total_high_value_lifetime = sum(c["amount"] for c in high_value_churn)
print(f"  Churned: {len(churned)}  high-value: {len(high_value_churn)}  RM {total_high_value_lifetime:,.0f}")

# ── Customer Insights (V1 第 3 刀): RFM by Date Enrol ────────────────
# Tier the member by membership-age bucket at snapshot; only count members
# with at least one purchase at any time, primary branch in active 5.
def age_bucket(years):
    if years < 1:  return "<1y"
    if years < 5:  return "1-5y"
    if years < 8:  return "5-8y"
    return "8y+"

# Members with valid enrol + active branch + any purchase
ci_rows = []
for mc, d in mem.items():
    if d["last"] is None: continue
    if not d["branches"]: continue
    primary_branch = max(d["branches"].items(), key=lambda x: x[1])[0]
    if primary_branch not in ACTIVE_BRANCHES: continue
    if d["enrol"] is None: continue
    years = (NOW_DT - d["enrol"]).days / 365.25
    if years < 0: continue
    bucket = age_bucket(years)
    ci_rows.append({
        "mc": str(mc),
        "name": (d["name"][:40] if d["name"] else str(mc)),
        "branch": primary_branch,
        "cust_type": d["cust_type"] or "Other",
        "enrol": d["enrol"].strftime("%Y-%m-%d"),
        "age_years": round(years, 1),
        "age_bucket": bucket,
        "ltm_amt": round(d["ltm_amt"], 0),
        "ltm_visits": len(d["ltm_visits"]),
        "m6_amt": round(d["m6_amt"], 0),
        "m6_visits": len(d["m6_visits"]),
        "m3_amt": round(d["m3_amt"], 0),
        "m3_visits": len(d["m3_visits"]),
        "m1_amt": round(d["m1_amt"], 0),
        "m1_visits": len(d["m1_visits"]),
        "lifetime_amt": round(d["amt"], 0),
        "last": d["last"].strftime("%Y-%m"),
    })
print(f"  CI members (enrol+active+purchased): {len(ci_rows):,}")

BUCKETS = ["<1y", "1-5y", "5-8y", "8y+"]
TYPES   = ["Walk-in", "Contractor", "Interior Designer", "Other"]
WINDOWS = ["1m", "3m", "6m", "12m"]   # most-recent-first selectable windows

def amt_field(w): return {"1m": "m1_amt", "3m": "m3_amt", "6m": "m6_amt", "12m": "ltm_amt"}[w]
def vis_field(w): return {"1m": "m1_visits", "3m": "m3_visits", "6m": "m6_visits", "12m": "ltm_visits"}[w]

# Per-bucket aggregate, per window — so the browser can switch windows w/o recomputing
bucket_agg_by_window = {}
for w in WINDOWS:
    af, vf = amt_field(w), vis_field(w)
    bagg = {b: {"n": 0, "amt": 0.0, "visits": 0, "n_repeat": 0, "n_active": 0} for b in BUCKETS}
    for r in ci_rows:
        b = bagg[r["age_bucket"]]
        b["n"] += 1
        b["amt"] += r[af]
        b["visits"] += r[vf]
        if r[vf] >= 1: b["n_active"] += 1
        if r[vf] >= 2: b["n_repeat"] += 1
    for b, v in bagg.items():
        v["aov"] = round(v["amt"] / v["visits"], 0) if v["visits"] else 0
        v["repeat_pct"] = round(100 * v["n_repeat"] / v["n_active"], 1) if v["n_active"] else 0
        v["amt"] = round(v["amt"], 0)
    bucket_agg_by_window[w] = bagg

# Cross-table type × bucket, per window
cross_by_window = {}
for w in WINDOWS:
    af = amt_field(w)
    cr = {tp: {b: {"n": 0, "amt": 0.0} for b in BUCKETS} for tp in TYPES}
    for r in ci_rows:
        tp = r["cust_type"] if r["cust_type"] in TYPES else "Other"
        cell = cr[tp][r["age_bucket"]]
        cell["n"] += 1
        cell["amt"] += r[af]
    for tp in TYPES:
        for b in BUCKETS:
            cr[tp][b]["amt"] = round(cr[tp][b]["amt"], 0)
    cross_by_window[w] = cr

# Top 100 default by LTM (browser re-sorts when window changes)
ci_rows_top = sorted(ci_rows, key=lambda x: -x["ltm_amt"])[:100]

# Headline numbers per window
def summary_for(w):
    af = amt_field(w)
    bagg = bucket_agg_by_window[w]
    total_n = len(ci_rows)
    n_5plus = bagg["5-8y"]["n"] + bagg["8y+"]["n"]
    total_amt = sum(r[af] for r in ci_rows)
    amt_5plus = sum(r[af] for r in ci_rows if r["age_bucket"] in ("5-8y", "8y+"))
    n_active  = sum(1 for r in ci_rows if r[af] > 0)
    return {
        "total_members": total_n,
        "n_active": n_active,
        "n_lt1":   bagg["<1y"]["n"],
        "n_1_5":   bagg["1-5y"]["n"],
        "n_5_8":   bagg["5-8y"]["n"],
        "n_8plus": bagg["8y+"]["n"],
        "amt_total": round(total_amt, 0),
        "amt_lt1":   bagg["<1y"]["amt"],
        "amt_1_5":   bagg["1-5y"]["amt"],
        "amt_5_8":   bagg["5-8y"]["amt"],
        "amt_8plus": bagg["8y+"]["amt"],
        "pct_5plus_n":   round(100 * n_5plus / total_n, 1) if total_n else 0,
        "pct_5plus_amt": round(100 * amt_5plus / total_amt, 1) if total_amt else 0,
    }

ci_summary_by_window = {w: summary_for(w) for w in WINDOWS}
ci_summary = ci_summary_by_window["12m"]   # default for legacy fields
ci_summary["snapshot"] = ym_str(NOW_YM)
print(f"  CI windows: " + " · ".join(
    f"{w}: {ci_summary_by_window[w]['n_active']:,} active / RM {ci_summary_by_window[w]['amt_total']:,.0f}"
    for w in WINDOWS))

# ── Procurement exceptions (C1) ──────────────────────────────────────
print("\nReading sales/stock/PO xlsx …")
wb = openpyxl.load_workbook(SVS_XLSX, read_only=True, data_only=True)
ws = wb["Raw Pivot"]
delayed = []   # closed PO, took >30 days
overdue = []   # open PO, no GRN, >30 days old
overdue_noise = 0
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or r[0] is None: continue
    pd, gd, code, br, pq, gq, pa, ga = r
    if not isinstance(pd, datetime): continue
    code = str(code).strip()
    pa = float(pa or 0)   # PO amount (committed)
    ga = float(ga or 0)   # GRN amount (received)
    pq = float(pq or 0)
    gq = float(gq or 0)
    if isinstance(gd, datetime):
        days = (gd - pd).days
        if days > 30:
            # closed PO — use GRN AMT (what actually arrived & was paid for)
            delayed.append({
                "code": code, "po_date": pd.strftime("%Y-%m-%d"),
                "grn_date": gd.strftime("%Y-%m-%d"),
                "days": days, "qty": gq, "amount": round(ga, 0),
                "kind": "delayed",
            })
    else:
        days = (NOW_DT - pd).days
        if days > 30:
            # open PO — use PO AMT (committed but not yet received)
            # filter noise: zero qty AND zero amt = cancelled stub, not real overdue
            if pq <= 0 and pa <= 0:
                overdue_noise += 1
                continue
            overdue.append({
                "code": code, "po_date": pd.strftime("%Y-%m-%d"),
                "grn_date": "", "days": days, "qty": pq, "amount": round(pa, 0),
                "kind": "overdue",
            })
wb.close()
print(f"  Overdue noise filtered (qty=0 & amt=0): {overdue_noise}")

# Pull SM brand metadata
print("Reading SM metadata …")
wb = openpyxl.load_workbook(SVS_XLSX, read_only=True, data_only=True)
ws = wb["SM"]
brand_map = {}
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or r[0] is None: continue
    code = str(r[0]).strip()
    brand = (r[10] or "") if isinstance(r[10], str) else ""
    sub = (r[5] or "") if isinstance(r[5], str) else ""
    mfr = (r[15] or "") if len(r) > 15 and isinstance(r[15], str) else ""
    brand_map[code] = {
        "brand": brand.strip(),
        "sub": sub.strip(),
        "manufacturer": mfr.strip(),
    }
wb.close()

for x in delayed + overdue:
    m = brand_map.get(x["code"], {})
    x["brand"] = m.get("brand", "")
    x["sub"] = m.get("sub", "")
    x["manufacturer"] = m.get("manufacturer", "")

# Sort: real overdue/delayed by amount desc (Jym wants to see real money first)
overdue.sort(key=lambda x: (-x["amount"], -x["days"]))
delayed.sort(key=lambda x: (-x["amount"], -x["days"]))
print(f"  Delayed POs (>30d, GRN AMT): {len(delayed)}  RM {sum(x['amount'] for x in delayed):,.0f}")
print(f"  Overdue open POs (>30d, PO AMT, noise-filtered): {len(overdue)}  RM {sum(x['amount'] for x in overdue):,.0f}")

# ── Sales 3m by branch (for B1 sales-drop signal) ────────────────────
print("\nComputing branch sales trends …")
wb = openpyxl.load_workbook(SVS_XLSX, read_only=True, data_only=True)
ws = wb["Raw sale"]
branch_month_amt = defaultdict(lambda: defaultdict(float))
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or r[0] is None: continue
    month, code, br, qty, amt = r
    if not isinstance(month, datetime): continue
    if br not in ACTIVE_BRANCHES: continue
    ym = (month.year, month.month)
    branch_month_amt[br][ym] += float(amt or 0)
wb.close()

LAST_3 = [shift(NOW_YM, n) for n in range(0, 3)]
PREV_3 = [shift(NOW_YM, n) for n in range(3, 6)]
branch_sales_trend = {}
for br in ACTIVE_BRANCHES:
    last = sum(branch_month_amt[br].get(ym, 0) for ym in LAST_3)
    prev = sum(branch_month_amt[br].get(ym, 0) for ym in PREV_3)
    drop_pct = round((1 - last / prev) * 100, 1) if prev > 0 else 0
    branch_sales_trend[br] = {
        "last_3m": round(last, 0),
        "prev_3m": round(prev, 0),
        "drop_pct": drop_pct,
    }

# ── Output: today-data.js ────────────────────────────────────────────
payload_today = {
    "meta": {
        "generated": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "snapshot": ym_str(NOW_YM),
    },
    "churn": {
        "summary": {
            "n_total": len(churned),
            "n_high_value": len(high_value_churn),
            "lifetime_rm": round(total_high_value_lifetime, 0),
            "cutoff_months": 6,
            "high_value_threshold": TIER_HIGH_THRESHOLD,
        },
        "rows": churned[:500],
    },
    "po_exceptions": {
        "summary": {
            "n_delayed": len(delayed),
            "n_overdue": len(overdue),
            "amount_delayed": round(sum(x["amount"] for x in delayed), 0),
            "amount_overdue": round(sum(x["amount"] for x in overdue), 0),
        },
        "delayed":  delayed[:300],
        "overdue":  overdue[:300],
    },
    "branch_sales_trend": branch_sales_trend,
}

os.makedirs(os.path.dirname(OUT_TODAY), exist_ok=True)
with open(OUT_TODAY, "w", encoding="utf-8") as f:
    f.write("/* AUTO-GENERATED by tools/build-today.py — DO NOT EDIT BY HAND */\n")
    f.write(f"/* Snapshot: {ym_str(NOW_YM)} */\n")
    f.write("window.WP_TODAY = ")
    json.dump(payload_today, f, ensure_ascii=False, separators=(",", ":"))
    f.write(";\n")
print(f"\nWrote {OUT_TODAY} ({os.path.getsize(OUT_TODAY):,} bytes)")

# ── Output: customers-data.js (V1 第 3 刀) ───────────────────────────
payload_cust = {
    "meta": {
        "generated": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "snapshot": ym_str(NOW_YM),
    },
    "summary": ci_summary,                           # legacy default (12m)
    "summary_by_window": ci_summary_by_window,
    "buckets_by_window": {
        w: [
            {
                "key": b,
                "n": bucket_agg_by_window[w][b]["n"],
                "amt": bucket_agg_by_window[w][b]["amt"],
                "aov": bucket_agg_by_window[w][b]["aov"],
                "repeat_pct": bucket_agg_by_window[w][b]["repeat_pct"],
                "n_active": bucket_agg_by_window[w][b]["n_active"],
            }
            for b in BUCKETS
        ]
        for w in WINDOWS
    },
    "cross_by_window": cross_by_window,
    "windows": WINDOWS,
    "types": TYPES,
    "top100": ci_rows_top,
    "all": ci_rows,           # for filter computations in browser
    # "race": removed in V1 第8刀 — now live via /api/floatation, not baked.
}

with open(OUT_CUST, "w", encoding="utf-8") as f:
    f.write("/* AUTO-GENERATED by tools/build-today.py — DO NOT EDIT BY HAND */\n")
    f.write(f"/* Snapshot: {ym_str(NOW_YM)} */\n")
    f.write("window.WP_CUSTOMERS = ")
    json.dump(payload_cust, f, ensure_ascii=False, separators=(",", ":"))
    f.write(";\n")
print(f"Wrote {OUT_CUST} ({os.path.getsize(OUT_CUST):,} bytes)")
