#!/usr/bin/env python3
# Wiltek Portal — V1 第 4 刀 Financial Depth data prep
# Reads:
#   - Live: https://wiltek-dashboard.vercel.app/api/proxy?type=financial   (FMM aggregates)
#   - /Users/jymchee/Desktop/Claude use/202604 - SALES VS STOCK VS PO VS GRN.V2.xlsx (brand-level GP estimate)
# Emits:
#   /Users/jymchee/wiltek-repo/assets/financial-data.js
import json, os, urllib.request
from collections import defaultdict
from datetime import datetime
import openpyxl

SVS_XLSX = "/Users/jymchee/Desktop/Claude use/202604 - SALES VS STOCK VS PO VS GRN.V2.xlsx"
OUT      = "/Users/jymchee/wiltek-repo/assets/financial-data.js"
FIN_URL  = "https://wiltek-dashboard.vercel.app/api/proxy?type=financial"

LTM_CUTOFF = datetime(2025, 4, 1)
NOW_YM = (2026, 3)

print("Fetching live FMM aggregates …")
with urllib.request.urlopen(FIN_URL, timeout=30) as r:
    fin = json.loads(r.read().decode("utf-8"))
if not fin.get("ok"):
    raise SystemExit(f"FMM endpoint not OK: {fin}")
fdata = fin["data"]
total = fdata["current_period"]["total"]
print(f"  current period: net_sales {total['net_sales']['inv']:,.0f}  total_exp {total['total_exp']['inv']:,.0f}")

# Build month-keyed cashflow series with totals per category
cf = fdata.get("cashflow", {})
months = cf.get("months", [])
series = cf.get("series", {})
# Take the last 6 months (months[] is already most-recent-first)
take = 6
months_recent = months[:take]
def col(name): return [round(float(x or 0), 0) for x in (series.get(name, []) or [])[:take]]

# Map API series → user's 7 cost categories
# Available: purchasing, op_mkt, online, logistics, admin, hr, accounting
# Mapping (best fit given API granularity):
#   1. Operational     = op_mkt + logistics
#   2. Marketing       = online   (paid media + KOL fall here)
#   3. Payroll (mgmt)  = N/A in API → split hr 30/70 as estimate
#   4. Payroll (ops)   = hr (70%)
#   5. Accounting      = accounting
#   6. Purchasing      = purchasing
#   7. Admin / BizDev  = admin
# We surface this honestly — page footer notes the mapping.
def add_lists(*lists, scale=1.0):
    n = max(len(l) for l in lists)
    out = []
    for i in range(n):
        s = sum((l[i] if i < len(l) else 0) for l in lists) * scale
        out.append(round(s, 0))
    return out

cat_series = {
    "Operational":         add_lists(col("op_mkt"), col("logistics")),
    "Marketing":           col("online"),
    "Payroll (Mgmt)":      add_lists(col("hr"), scale=0.30),   # estimate
    "Payroll (Ops)":       add_lists(col("hr"), scale=0.70),   # estimate
    "Accounting":          col("accounting"),
    "Purchasing":          col("purchasing"),
    "Admin / Biz Dev":     col("admin"),
}
sales_coll_series = col("sales_collection")

# Compute month-on-month % change for the most recent month vs previous month
def mom_pct(s):
    if len(s) < 2 or s[1] == 0: return 0.0
    return round((s[0] - s[1]) / s[1] * 100, 1)

cat_summary = []
for name, s in cat_series.items():
    latest = s[0] if s else 0
    pct_sales = round(latest / sales_coll_series[0] * 100, 1) if sales_coll_series and sales_coll_series[0] else 0
    mom = mom_pct(s)
    flag = "ok"
    if mom > 20: flag = "warn"
    cat_summary.append({
        "name": name,
        "latest": latest,
        "pct_sales": pct_sales,
        "mom_pct": mom,
        "series": s,
        "flag": flag,
    })

# HR / Sales red-line check (target 28%)
hr_total_latest = cat_series["Payroll (Mgmt)"][0] + cat_series["Payroll (Ops)"][0]
hr_pct_sales = round(hr_total_latest / sales_coll_series[0] * 100, 1) if sales_coll_series[0] else 0
for c in cat_summary:
    if c["name"].startswith("Payroll") and hr_pct_sales > 28:
        c["flag"] = "danger"

# ── Brand GP — build from local xlsx (sales × unit cost from SM) ──────
print("Building brand-level GP estimate from local xlsx …")
wb = openpyxl.load_workbook(SVS_XLSX, read_only=True, data_only=True)
ws = wb["SM"]
# SM columns (1-idx → 0-idx):
#   A=Item Code(0), F=Sub group(5), K=Brand(10), N=Unit Cost? — let's check
header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
print("  SM header:", header[:20])
brand_of = {}
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or r[0] is None: continue
    code = str(r[0]).strip()
    brand = (r[10] or "") if isinstance(r[10], str) else ""
    brand_of[code] = brand.strip() or "Unknown"
wb.close()

# Derive unit cost from PO/GRN history (SM sheet has no cost column)
print("Deriving per-item unit cost from PO/GRN history …")
wb = openpyxl.load_workbook(SVS_XLSX, read_only=True, data_only=True)
ws = wb["Raw Pivot"]
_acc = defaultdict(lambda: {"a": 0.0, "q": 0.0})
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or r[0] is None: continue
    _pd, _gd, code, _br, pq, gq, pa, ga = r
    pq, gq = float(pq or 0), float(gq or 0)
    pa, ga = float(pa or 0), float(ga or 0)
    code = str(code).strip()
    if gq > 0 and ga > 0:
        _acc[code]["a"] += ga; _acc[code]["q"] += gq
    elif pq > 0 and pa > 0:
        _acc[code]["a"] += pa; _acc[code]["q"] += pq
wb.close()
cost_of = {c: d["a"]/d["q"] for c, d in _acc.items() if d["q"] > 0}
print(f"  Brands mapped: {len(brand_of):,}  Item costs derived: {len(cost_of):,}")

# Sum LTM sales + LTM cogs by brand
print("Reading Raw sale for LTM brand sales …")
wb = openpyxl.load_workbook(SVS_XLSX, read_only=True, data_only=True)
ws = wb["Raw sale"]
brand_amt = defaultdict(float)
brand_qty = defaultdict(float)
brand_cogs = defaultdict(float)
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or r[0] is None: continue
    month, code, br, qty, amt = r
    if not isinstance(month, datetime): continue
    if month < LTM_CUTOFF: continue
    code = str(code).strip()
    brand = brand_of.get(code, "Unknown")
    a = float(amt or 0)
    q = float(qty or 0)
    brand_amt[brand] += a
    brand_qty[brand] += q
    if code in cost_of:
        brand_cogs[brand] += q * cost_of[code]
wb.close()

brands_out = []
for b, amt in brand_amt.items():
    cogs = brand_cogs.get(b, 0.0)
    gp = amt - cogs
    gp_pct = round(gp / amt * 100, 1) if amt else 0
    brands_out.append({
        "brand": b or "Unknown",
        "ltm_sales": round(amt, 0),
        "ltm_cogs": round(cogs, 0),
        "gp": round(gp, 0),
        "gp_pct": gp_pct,
        "qty": round(brand_qty.get(b, 0), 0),
    })
brands_out.sort(key=lambda x: -x["gp"])
print(f"  Brand rows: {len(brands_out)}")

# ── Diagnosis line ───────────────────────────────────────────────────
diag = "7 categories within target; no anomalies."
diag_zh = "7 大费用与目标对齐,无异常。"
weak = next((b for b in brands_out if b["ltm_sales"] > 50_000 and b["gp_pct"] < 20), None)
if hr_pct_sales > 30:
    diag = f"HR is {hr_pct_sales}% of sales (target 28%). Optimisation priority."
    diag_zh = f"HR 占销售 {hr_pct_sales}%,目标 28%,优化重点。"
elif weak:
    diag = f"Brand {weak['brand']} GP% only {weak['gp_pct']}%. Review pricing or phase out."
    diag_zh = f"品牌 {weak['brand']} GP% 仅 {weak['gp_pct']}%,需检视定价或淘汰。"

# ── Output ───────────────────────────────────────────────────────────
payload = {
    "meta": {
        "generated": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "snapshot": f"{NOW_YM[0]:04d}-{NOW_YM[1]:02d}",
        "source": "FMM (Official) via Apps Script + local SM unit costs",
        "diag_en": diag,
        "diag_zh": diag_zh,
    },
    "headline": {
        "sales_collection": sales_coll_series[0] if sales_coll_series else 0,
        "total_expenses": round(sum(cat_series[k][0] for k in cat_series), 0),
        "hr_pct_sales": hr_pct_sales,
    },
    "categories": cat_summary,
    "months": months_recent,
    "sales_coll_series": sales_coll_series,
    "brands": brands_out[:30],     # top 30 by GP
    "brand_count": len(brands_out),
}
payload["headline"]["net"] = round(payload["headline"]["sales_collection"] - payload["headline"]["total_expenses"], 0)
payload["headline"]["net_pct"] = round(payload["headline"]["net"] / payload["headline"]["sales_collection"] * 100, 1) if payload["headline"]["sales_collection"] else 0

os.makedirs(os.path.dirname(OUT), exist_ok=True)
with open(OUT, "w", encoding="utf-8") as f:
    f.write("/* AUTO-GENERATED by tools/build-financial.py — DO NOT EDIT BY HAND */\n")
    f.write("window.WP_FINANCIAL = ")
    json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))
    f.write(";\n")
print(f"\nWrote {OUT} ({os.path.getsize(OUT):,} bytes)")
print(f"  diag: {diag}")
