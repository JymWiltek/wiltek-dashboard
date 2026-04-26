#!/usr/bin/env python3
# ─────────────────────────────────────────────────────────────────────
# Wiltek Portal — BI Cut #2 data prep
# Reads:  Desktop/Claude use/202604 - SALES VS STOCK VS PO VS GRN.V2.xlsx
# Emits:  /Users/jymchee/wiltek-repo/assets/sku-tiers.js
# Shape:  window.WP_SKU = { meta, tiers, watch, matrix, matrix_diag,
#                            ampang, recommended_a, recommended_b, avoid_d }
#
# Tiering: ABC by 6-month sales (cumulative 80/95/99). D-overlay if any:
#   - 60+ day zero sales (last 2 months sold = 0)
#   - stock ÷ monthly avg sales > 6
#   - 3-month sales ≥ 50% drop vs prior 3 months
# Period: latest 6 months (Oct 2025 → Mar 2026), 6 ops branches only.
# ─────────────────────────────────────────────────────────────────────
import json
import os
from collections import defaultdict
from datetime import datetime
import openpyxl

XLSX = "/Users/jymchee/Desktop/Claude use/202604 - SALES VS STOCK VS PO VS GRN.V2.xlsx"
OUT  = "/Users/jymchee/wiltek-repo/assets/sku-tiers.js"

# 6 operational branches only (matches WP_PERMS.ALL_BRANCHES)
BRANCHES = ["W01", "W02", "W03", "W05", "W07", "W11"]

# Latest 6 months of sales we trust. Today is 2026-04-26 — March is the last
# fully closed month in the workbook. Window: Oct 2025 → Mar 2026 (6 months).
P6_MONTHS  = [(2025,10),(2025,11),(2025,12),(2026,1),(2026,2),(2026,3)]
P3_LATE    = [(2026,1),(2026,2),(2026,3)]
P3_EARLY   = [(2025,10),(2025,11),(2025,12)]
P2_LATEST  = [(2026,2),(2026,3)]   # used for "60-day zero sales"

AMPANG_OPEN  = (2026, 5, 1)
TODAY        = (2026, 4, 26)

# ── Load workbook ───────────────────────────────────────────────────
print(f"Loading {XLSX} …")
wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

# ── Raw sale → per-(SKU,branch,month) qty + amount ──────────────────
ws = wb["Raw sale"]
# (sku, branch, (y,m)) -> {qty, amt}
sales: dict = defaultdict(lambda: {"qty": 0.0, "amt": 0.0})
# (sku, (y,m)) total
sku_month: dict = defaultdict(float)
# sku totals over 6m
sku_6m_amt: dict = defaultdict(float)
sku_6m_qty: dict = defaultdict(float)
# (sku, branch) 6m amt
sku_branch_6m_amt: dict = defaultdict(float)

rows = ws.iter_rows(min_row=2, values_only=True)
for r in rows:
    if not r or r[0] is None: continue
    month, code, branch, qty, amt = r[0], r[1], r[2], r[3], r[4]
    if branch not in BRANCHES: continue
    if not isinstance(month, datetime): continue
    ym = (month.year, month.month)
    if ym not in P6_MONTHS: continue
    code = str(code).strip()
    qty = float(qty or 0)
    amt = float(amt or 0)
    sales[(code, branch, ym)]["qty"] += qty
    sales[(code, branch, ym)]["amt"] += amt
    sku_month[(code, ym)] += amt
    sku_6m_amt[code] += amt
    sku_6m_qty[code] += qty
    sku_branch_6m_amt[(code, branch)] += amt

print(f"Sales rows mapped to 6m × 6 branches: {len(sales)}")
print(f"Unique SKUs with any 6m sales: {len(sku_6m_amt)}")

# ── Raw CS → current stock per (sku, branch) ────────────────────────
ws = wb["Raw CS"]
sku_stock_qty: dict = defaultdict(float)         # total qty across branches
sku_stock_val: dict = defaultdict(float)         # total $ value
branch_sku_stock_qty: dict = defaultdict(float)  # (sku, branch) qty
branch_sku_stock_val: dict = defaultdict(float)
total_stock_val = 0.0

rows = ws.iter_rows(min_row=2, values_only=True)
for r in rows:
    if not r or r[0] is None: continue
    code, branch, qty, unit_cost, on_hand = r[0], r[1], r[2], r[3], r[4]
    if branch not in BRANCHES: continue
    code = str(code).strip()
    qty = float(qty or 0)
    on_hand = float(on_hand or 0)  # already qty*unit_cost
    sku_stock_qty[code] += qty
    sku_stock_val[code] += on_hand
    branch_sku_stock_qty[(code, branch)] += qty
    branch_sku_stock_val[(code, branch)] += on_hand
    total_stock_val += on_hand

print(f"Total stock value (6 branches): RM {total_stock_val:,.0f}")

# ── SM → metadata (name/category/brand) ─────────────────────────────
ws = wb["SM"]
sku_meta: dict = {}
rows = ws.iter_rows(min_row=2, values_only=True)
for r in rows:
    if not r or r[0] is None: continue
    code = str(r[0]).strip()
    sku_meta[code] = {
        "department": (r[3] or "").strip() if isinstance(r[3], str) else "",
        "main_group": (r[4] or "").strip() if isinstance(r[4], str) else "",
        "sub_group":  (r[5] or "").strip() if isinstance(r[5], str) else "",
        "brand":      (r[10] or "").strip() if isinstance(r[10], str) else "",
        "movement":   (r[13] or "").strip() if isinstance(r[13], str) else "",
        "status":     (r[2] or "").strip() if isinstance(r[2], str) else "",
    }

print(f"SM metadata rows: {len(sku_meta)}")

# ── Universe of SKUs we care about ──────────────────────────────────
all_skus = set(sku_6m_amt.keys()) | set(sku_stock_qty.keys())
# Discard discontinued/obsolete that have neither sales nor stock
all_skus = {s for s in all_skus if sku_6m_amt.get(s, 0) > 0 or sku_stock_qty.get(s, 0) > 0}
print(f"SKU universe: {len(all_skus)}")

total_6m_amt = sum(sku_6m_amt.values())
print(f"Total 6m sales: RM {total_6m_amt:,.0f}")

# ── ABC tiering by 6m sales (cum 80/95/99) ──────────────────────────
sorted_skus = sorted(all_skus, key=lambda s: sku_6m_amt.get(s, 0), reverse=True)
tier: dict = {}
cum = 0.0
for s in sorted_skus:
    sales_amt = sku_6m_amt.get(s, 0)
    if total_6m_amt > 0:
        cum += sales_amt
        pct = cum / total_6m_amt
    else:
        pct = 1.0
    if sales_amt <= 0:
        tier[s] = "D"  # zero sales over 6m → dead by definition
    elif pct <= 0.80:
        tier[s] = "A"
    elif pct <= 0.95:
        tier[s] = "B"
    elif pct <= 0.99:
        tier[s] = "C"
    else:
        tier[s] = "D"

# ── D overlay (any condition pulls a sku to D) ──────────────────────
def is_dead(s: str) -> tuple[bool, str]:
    sales_6m = sku_6m_amt.get(s, 0)
    stock_qty = sku_stock_qty.get(s, 0)
    stock_val = sku_stock_val.get(s, 0)
    # 60-day zero
    last2 = sum(sku_month.get((s, ym), 0) for ym in P2_LATEST)
    if stock_qty > 0 and last2 == 0:
        return True, f"60 天零销售,库存 RM {stock_val:,.0f}"
    # stock cover > 6 months
    monthly = sales_6m / 6.0 if sales_6m > 0 else 0
    if monthly > 0 and (stock_val / monthly) > 6:
        cover_m = stock_val / monthly
        return True, f"库存够卖 {cover_m:.1f} 个月(>6)"
    if monthly == 0 and stock_val > 0:
        return True, f"库存 RM {stock_val:,.0f},零销售"
    # 3-month decline ≥ 50%
    p3l = sum(sku_month.get((s, ym), 0) for ym in P3_LATE)
    p3e = sum(sku_month.get((s, ym), 0) for ym in P3_EARLY)
    if p3e >= 1000 and p3l <= p3e * 0.5:
        drop = (1 - (p3l / p3e)) * 100 if p3e > 0 else 0
        return True, f"近 3 月销售比上 3 月跌 {drop:.0f}%"
    return False, ""

dead_reason: dict = {}
for s in list(all_skus):
    dead, reason = is_dead(s)
    if dead:
        tier[s] = "D"
        dead_reason[s] = reason

# Tier rollup (4 cards)
tiers_data = []
for t in ["A", "B", "C", "D"]:
    members = [s for s in all_skus if tier.get(s) == t]
    cnt = len(members)
    sales = sum(sku_6m_amt.get(s, 0) for s in members)
    stock_v = sum(sku_stock_val.get(s, 0) for s in members)
    pct_sales = (sales / total_6m_amt * 100) if total_6m_amt > 0 else 0
    pct_stock = (stock_v / total_stock_val * 100) if total_stock_val > 0 else 0
    tiers_data.append({
        "tier": t,
        "count": cnt,
        "sales_6m": round(sales),
        "stock_val": round(stock_v),
        "pct_sales": round(pct_sales, 1),
        "pct_stock": round(pct_stock, 1),
    })

# Auto-diagnostic for D card
d = next(t for t in tiers_data if t["tier"] == "D")
d_diag_zh = (
    f"D 级呆货 {d['count']} 个 SKU,占库存 {d['pct_stock']:.0f}%,"
    f"6 个月只贡献 {d['pct_sales']:.1f}% 销售。建议清仓回笼现金。"
)
d_diag_en = (
    f"{d['count']} D-tier dead SKUs hold {d['pct_stock']:.0f}% of stock value but only "
    f"contribute {d['pct_sales']:.1f}% of 6m sales. Liquidate to recover cash."
)

a = next(t for t in tiers_data if t["tier"] == "A")
a_diag_zh = f"A 级 {a['count']} 个 SKU 撑起 {a['pct_sales']:.0f}% 销售,确保 0 缺货。"
a_diag_en = f"{a['count']} A-tier SKUs drive {a['pct_sales']:.0f}% of sales — guarantee 0 stockouts."
b = next(t for t in tiers_data if t["tier"] == "B")
b_diag_zh = f"B 级 {b['count']} 个 SKU,稳定补货,避免断货。"
b_diag_en = f"{b['count']} B-tier SKUs — keep replenished, avoid stockouts."
c = next(t for t in tiers_data if t["tier"] == "C")
c_diag_zh = f"C 级 {c['count']} 个 SKU,按需采购,不积压。"
c_diag_en = f"{c['count']} C-tier SKUs — order on demand, do not stockpile."

a["diag_zh"], a["diag_en"] = a_diag_zh, a_diag_en
b["diag_zh"], b["diag_en"] = b_diag_zh, b_diag_en
c["diag_zh"], c["diag_en"] = c_diag_zh, c_diag_en
d["diag_zh"], d["diag_en"] = d_diag_zh, d_diag_en

# ── Top 5 SKUs to watch — risk × value heuristic ───────────────────
def risk_score(s: str) -> tuple[float, str, str]:
    """Returns (score, reason_zh, reason_en). Higher = more urgent."""
    sales_6m = sku_6m_amt.get(s, 0)
    stock_v  = sku_stock_val.get(s, 0)
    stock_q  = sku_stock_qty.get(s, 0)
    monthly = sales_6m / 6.0 if sales_6m > 0 else 0
    p3l = sum(sku_month.get((s, ym), 0) for ym in P3_LATE)
    p3e = sum(sku_month.get((s, ym), 0) for ym in P3_EARLY)
    last2 = sum(sku_month.get((s, ym), 0) for ym in P2_LATEST)

    # Pattern 1: high stock + zero recent sales (worst dead stock)
    if stock_v >= 5000 and last2 == 0:
        return (stock_v * 1.5,
                f"库存 RM {stock_v:,.0f},近 2 个月 0 销售",
                f"Stock RM {stock_v:,.0f}, 0 sales in last 2 months")
    # Pattern 2: hot SKU with low stock cover (stockout risk)
    if monthly >= 3000 and stock_v < monthly * 1.0:
        cover = (stock_v / monthly) if monthly > 0 else 0
        return (monthly * 2.0,
                f"月销 RM {monthly:,.0f},库存只够卖 {cover:.1f} 个月,断货风险",
                f"Monthly RM {monthly:,.0f}, only {cover:.1f}-month cover — stockout risk")
    # Pattern 3: declining ≥50% with meaningful base
    if p3e >= 5000 and p3l <= p3e * 0.5:
        drop = (1 - (p3l / p3e)) * 100 if p3e > 0 else 0
        return (p3e * (drop / 100) * 1.2,
                f"近 3 月销售比上 3 月跌 {drop:.0f}%(RM {p3e:,.0f}→{p3l:,.0f})",
                f"Last 3m vs prior 3m fell {drop:.0f}% (RM {p3e:,.0f}→{p3l:,.0f})")
    # Pattern 4: dead w/ medium stock
    if stock_v >= 3000 and last2 == 0 and sales_6m == 0:
        return (stock_v * 1.2,
                f"6 个月零销售,库存 RM {stock_v:,.0f}",
                f"6m zero sales, stock RM {stock_v:,.0f}")
    return (0, "", "")

ranked = []
for s in all_skus:
    score, rzh, ren = risk_score(s)
    if score > 0:
        ranked.append((score, s, rzh, ren))
ranked.sort(reverse=True)

watch_data = []
for score, s, rzh, ren in ranked[:5]:
    meta = sku_meta.get(s, {})
    name = (meta.get("main_group") or "") + (" · " + meta.get("sub_group") if meta.get("sub_group") else "")
    name = name.strip(" ·") or s
    if meta.get("brand"):
        name = f"{meta['brand']} {name}"
    sales_6m = sku_6m_amt.get(s, 0)
    stock_v  = sku_stock_val.get(s, 0)
    p3l = sum(sku_month.get((s, ym), 0) for ym in P3_LATE)
    p3e = sum(sku_month.get((s, ym), 0) for ym in P3_EARLY)
    watch_data.append({
        "code": s,
        "name": name[:60],
        "tier": tier.get(s, "D"),
        "sales_6m": round(sales_6m),
        "stock_val": round(stock_v),
        "p3_late": round(p3l),
        "p3_early": round(p3e),
        "diag_zh": rzh,
        "diag_en": ren,
    })

# ── 6×4 matrix (branch × tier) ──────────────────────────────────────
matrix = {}  # branch -> tier -> {cnt, sales_pct, stock_pct, sales_amt, stock_val}
branch_total_sales = defaultdict(float)
branch_total_stock = defaultdict(float)
for s in all_skus:
    for b in BRANCHES:
        branch_total_sales[b] += sku_branch_6m_amt.get((s, b), 0)
        branch_total_stock[b] += branch_sku_stock_val.get((s, b), 0)

for b in BRANCHES:
    matrix[b] = {}
    for t in ["A", "B", "C", "D"]:
        members = [s for s in all_skus if tier.get(s) == t]
        # SKUs that have any presence at branch b (sales OR stock)
        present = [s for s in members
                   if sku_branch_6m_amt.get((s, b), 0) > 0
                   or branch_sku_stock_qty.get((s, b), 0) > 0]
        sales_amt = sum(sku_branch_6m_amt.get((s, b), 0) for s in present)
        stock_v   = sum(branch_sku_stock_val.get((s, b), 0) for s in present)
        sp = (sales_amt / branch_total_sales[b] * 100) if branch_total_sales[b] > 0 else 0
        kp = (stock_v / branch_total_stock[b] * 100) if branch_total_stock[b] > 0 else 0
        matrix[b][t] = {
            "count": len(present),
            "sales_pct": round(sp, 1),
            "stock_pct": round(kp, 1),
            "sales_amt": round(sales_amt),
            "stock_val": round(stock_v),
        }

# Matrix diagnostic — flag the worst offender (highest D stock_pct)
worst_b, worst_pct = None, 0
for b in BRANCHES:
    p = matrix[b]["D"]["stock_pct"]
    if p > worst_pct:
        worst_pct, worst_b = p, b
matrix_diag_zh = ""
matrix_diag_en = ""
if worst_b and worst_pct >= 20:
    matrix_diag_zh = f"{worst_b} D 级 SKU 占库存 {worst_pct:.0f}%,该店清仓优先级最高。"
    matrix_diag_en = f"{worst_b} has {worst_pct:.0f}% stock value tied up in D-tier — top liquidation priority."
elif worst_b:
    matrix_diag_zh = f"各店 D 级占比均 ≤ {worst_pct:.0f}%,呆货分布相对均匀。"
    matrix_diag_en = f"All branches keep D-tier ≤ {worst_pct:.0f}% — dead stock spread evenly."

# ── Ampang banner (5/1 opening) ────────────────────────────────────
def days_between(a, b):
    da = datetime(*a)
    db = datetime(*b)
    return (db - da).days

days_to_ampang = days_between(TODAY, AMPANG_OPEN)

a_skus = [s for s in all_skus if tier.get(s) == "A"]
b_skus = [s for s in all_skus if tier.get(s) == "B"]
d_skus = [s for s in all_skus if tier.get(s) == "D"]

# Top 240 A by 6m sales (limited by available)
recommended_a = sorted(a_skus, key=lambda s: sku_6m_amt.get(s, 0), reverse=True)[:240]
# Top 100 B (front-runners) by 6m sales
b_front = sorted(b_skus, key=lambda s: sku_6m_amt.get(s, 0), reverse=True)[:100]
avoid_d = sorted(d_skus, key=lambda s: sku_stock_val.get(s, 0), reverse=True)[:380]

ampang = {
    "days": days_to_ampang,
    "open_date": "2026-05-01",
    "rec_a_count": len(recommended_a),
    "rec_b_count": len(b_front),
    "avoid_d_count": len(avoid_d),
    "rec_a_value": round(sum(sku_6m_amt.get(s, 0) for s in recommended_a)),
    "avoid_d_value": round(sum(sku_stock_val.get(s, 0) for s in avoid_d)),
}

# ── Compose payload ────────────────────────────────────────────────
payload = {
    "meta": {
        "generated": "2026-04-26",
        "period": "2025-10 → 2026-03 (6 months)",
        "branches": BRANCHES,
        "total_skus": len(all_skus),
        "total_sales_6m": round(total_6m_amt),
        "total_stock_val": round(total_stock_val),
    },
    "tiers": tiers_data,
    "watch": watch_data,
    "matrix": matrix,
    "matrix_diag_zh": matrix_diag_zh,
    "matrix_diag_en": matrix_diag_en,
    "ampang": ampang,
    # Lists are kept short — UI doesn't render them, only the count is shown.
    # Keeping ids out of the payload makes the JS file ~10 KB instead of 100+.
}

# ── Emit JS ─────────────────────────────────────────────────────────
os.makedirs(os.path.dirname(OUT), exist_ok=True)
with open(OUT, "w", encoding="utf-8") as f:
    f.write("/* AUTO-GENERATED by tools/build-sku-tiers.py — DO NOT EDIT BY HAND */\n")
    f.write("/* Source: 202604 - SALES VS STOCK VS PO VS GRN.V2.xlsx */\n")
    f.write("/* Period: 2025-10 → 2026-03 · 6 ops branches (W01/02/03/05/07/11) */\n")
    f.write("window.WP_SKU = ")
    f.write(json.dumps(payload, ensure_ascii=False, indent=2))
    f.write(";\n")

print(f"\nWrote {OUT}")
print(f"  tiers     : {[ (t['tier'], t['count']) for t in tiers_data ]}")
print(f"  watch     : {len(watch_data)} SKUs")
print(f"  matrix    : 6 × 4 cells")
print(f"  ampang    : {days_to_ampang} days, A={ampang['rec_a_count']} B={ampang['rec_b_count']} D={ampang['avoid_d_count']}")
