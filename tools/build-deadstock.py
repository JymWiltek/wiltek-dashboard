#!/usr/bin/env python3
# Wiltek Portal — V1 第 0 刀 data prep
# Reads:  /Users/jymchee/Desktop/Claude use/202604 - SALES VS STOCK VS PO VS GRN.V2.xlsx
# Emits:  /Users/jymchee/wiltek-repo/assets/deadstock-data.js
#
# Classifies every stock row in the 5 active branches into one of:
#   ACTIVE        — sold in this branch within last 6 months
#   SLOW          — last sale in this branch was 6–12 months ago
#   DEAD          — last sale in this branch was 12+ months ago
#   MISPLACED     — never sold in this branch, but other active branches sold it in last 6 months
#   COMPANY_DEAD  — never sold anywhere active in last 6 months and never in this branch
import json
import os
from collections import defaultdict
from datetime import datetime
import openpyxl

XLSX = "/Users/jymchee/Desktop/Claude use/202604 - SALES VS STOCK VS PO VS GRN.V2.xlsx"
OUT  = "/Users/jymchee/wiltek-repo/assets/deadstock-data.js"

ACTIVE = ["W01", "W02", "W03", "W05", "W07"]
BRANCH_NAMES = {
    "W01": "Pandan Indah",
    "W02": "Ampang Waterfront",
    "W03": "Wangsa Maju",
    "W05": "Bangi Seksyen 7",
    "W07": "Pandan Jaya",
}

def shift(ym, n):
    y, m = ym
    total = y * 12 + (m - 1) - n
    return (total // 12, total % 12 + 1)

def ym_str(ym):
    return f"{ym[0]:04d}-{ym[1]:02d}"

print(f"Loading {XLSX} …")
wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

# ── SM metadata (description / category / brand) ─────────────────────
ws = wb["SM"]
meta = {}
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or r[0] is None: continue
    code = str(r[0]).strip()
    department = (r[3] or "") if isinstance(r[3], str) else ""
    main_group = (r[4] or "") if isinstance(r[4], str) else ""
    sub_group  = (r[5] or "") if isinstance(r[5], str) else ""
    brand      = (r[10] or "") if isinstance(r[10], str) else ""
    meta[code] = {
        "department": department.strip(),
        "category":   main_group.strip(),
        "sub":        sub_group.strip(),
        "brand":      brand.strip(),
    }

# ── Raw sale ────────────────────────────────────────────────────────
ws = wb["Raw sale"]
all_rows = []
max_month = None
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or r[0] is None: continue
    month, code, branch, qty, amt = r[0], r[1], r[2], r[3], r[4]
    if not isinstance(month, datetime): continue
    code = str(code).strip()
    branch = str(branch).strip()
    qty = float(qty or 0)
    ym = (month.year, month.month)
    if max_month is None or ym > max_month:
        max_month = ym
    all_rows.append((ym, code, branch, qty))

NOW = max_month                # latest month with data, e.g. (2026, 3)
SLOW_CUTOFF = shift(NOW, 5)    # earliest month considered "recent" (last 6m incl.)
DEAD_CUTOFF = shift(NOW, 11)   # earliest month considered "still warm" (last 12m incl.)
SALES_3M_CUTOFF = shift(NOW, 2)  # earliest month for last-3m sales (used for transfer demand)
print(f"Snapshot month: {ym_str(NOW)}")
print(f"SLOW cutoff (last 6m incl.): {ym_str(SLOW_CUTOFF)}")
print(f"DEAD cutoff (last 12m incl.): {ym_str(DEAD_CUTOFF)}")
print(f"Sales-3m cutoff (transfer demand): {ym_str(SALES_3M_CUTOFF)}")

# Aggregations
last_sale_skubr = {}                                        # (sku, branch) -> last ym (any branch, but keyed per pair)
sku_branch_qty_last6  = defaultdict(lambda: defaultdict(float))  # sku -> branch -> qty in last 6m (active branches only)
sku_branch_qty_last3  = defaultdict(lambda: defaultdict(float))  # sku -> branch -> qty in last 3m (active branches only)
sku_branch_qty_last12 = defaultdict(lambda: defaultdict(float))  # sku -> branch -> qty in last 12m (active branches only)

for (ym, code, branch, qty) in all_rows:
    key = (code, branch)
    if key not in last_sale_skubr or ym > last_sale_skubr[key]:
        last_sale_skubr[key] = ym
    if branch in ACTIVE and ym >= SLOW_CUTOFF:
        sku_branch_qty_last6[code][branch] += qty
    if branch in ACTIVE and ym >= SALES_3M_CUTOFF:
        sku_branch_qty_last3[code][branch] += qty
    if branch in ACTIVE and ym >= DEAD_CUTOFF:
        sku_branch_qty_last12[code][branch] += qty

# ── Raw CS (current stock) ──────────────────────────────────────────
# Two passes through the same sheet:
#   pass A — count EVERY row across ALL locations to compute company-wide
#            total stock value (5 active stores + WLO/W11/WL1/WSR warehouse
#            depots etc.). The Stock main dashboard surfaces this so the
#            "总库存值" reflects the whole business, not just shelf displays.
#   pass B — only the 5 active stores get classified into ACTIVE/SLOW/DEAD/
#            MISPLACED/COMPANY_DEAD because dead-stock classification needs
#            sales history per branch — warehouse depots have no sales by
#            design.
ws = wb["Raw CS"]
company_total_stock = 0.0
company_rows_total  = 0
company_by_branch   = {}   # branch -> on_hand
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or r[0] is None: continue
    code, branch, qty, unit_cost, on_hand = r[0], r[1], r[2], r[3], r[4]
    branch = str(branch).strip() if branch else ''
    on_hand = float(on_hand or 0)
    company_total_stock += on_hand
    company_rows_total  += 1
    company_by_branch[branch] = company_by_branch.get(branch, 0.0) + on_hand

ws = wb["Raw CS"]
classified = []  # list of dicts
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or r[0] is None: continue
    code, branch, qty, unit_cost, on_hand = r[0], r[1], r[2], r[3], r[4]
    if branch not in ACTIVE: continue
    code = str(code).strip()
    qty = float(qty or 0)
    unit_cost = float(unit_cost or 0)
    on_hand = float(on_hand or 0)

    last = last_sale_skubr.get((code, branch))
    others_qty = sum(q for b, q in sku_branch_qty_last6[code].items() if b != branch and b in ACTIVE)

    if last and last >= SLOW_CUTOFF:
        cls = "ACTIVE"
    elif last and last >= DEAD_CUTOFF:
        cls = "SLOW"
    elif last:
        cls = "DEAD"
    elif others_qty > 0:
        cls = "MISPLACED"
    else:
        cls = "COMPANY_DEAD"

    m = meta.get(code, {})
    classified.append({
        "cls":        cls,
        "branch":     branch,
        "code":       code,
        "desc":       m.get("brand", "") + ((" · " + m.get("sub", "")) if m.get("sub") else ""),
        "category":   m.get("category", "") or "—",
        "department": m.get("department", "") or "",
        "qty":        round(qty, 2),
        "unit_cost":  round(unit_cost, 2),
        "amount":     round(on_hand, 2),
        "last_sale":  ym_str(last) if last else "",
        "others_qty": round(others_qty, 2),
    })

# Per-branch: build distribution of this SKU across active branches (for row-expand panel)
# (sku -> {branch: qty_in_stock})
sku_branch_stock = defaultdict(lambda: defaultdict(float))
for row in classified:
    sku_branch_stock[row["code"]][row["branch"]] = row["qty"]

# ── Aggregate totals & per-branch ───────────────────────────────────
totals = defaultdict(lambda: {"rows": 0, "amount": 0.0})
by_branch = {b: {
    "name": BRANCH_NAMES[b],
    "total": 0.0,
    "ACTIVE": 0.0, "SLOW": 0.0, "DEAD": 0.0, "MISPLACED": 0.0, "COMPANY_DEAD": 0.0,
    "rows": 0,
} for b in ACTIVE}

for row in classified:
    totals[row["cls"]]["rows"] += 1
    totals[row["cls"]]["amount"] += row["amount"]
    b = row["branch"]
    by_branch[b]["total"] += row["amount"]
    by_branch[b][row["cls"]] += row["amount"]
    by_branch[b]["rows"] += 1

# Compute problem percentage per branch (everything except ACTIVE)
for b, info in by_branch.items():
    problem = info["total"] - info["ACTIVE"]
    info["problem"] = round(problem, 2)
    info["problem_pct"] = round((problem / info["total"] * 100) if info["total"] > 0 else 0, 1)
    for k in ("total", "ACTIVE", "SLOW", "DEAD", "MISPLACED", "COMPANY_DEAD"):
        info[k] = round(info[k], 2)

total_stock = sum(t["amount"] for t in totals.values())
problem_total = total_stock - totals["ACTIVE"]["amount"]

# Summary print (sanity)
print("\n--- Computed totals (5 active branches) ---")
for k in ["ACTIVE", "SLOW", "DEAD", "MISPLACED", "COMPANY_DEAD"]:
    t = totals[k]
    print(f"  {k:<14} {t['rows']:>6}  RM {t['amount']:>12,.0f}")
print(f"  {'TOTAL':<14} {sum(t['rows'] for t in totals.values()):>6}  RM {total_stock:>12,.0f}")
print(f"  {'PROBLEM':<14} {'':>6}  RM {problem_total:>12,.0f}  ({problem_total/total_stock*100:.1f}%)")

# Round amounts in totals
totals_clean = {k: {"rows": v["rows"], "amount": round(v["amount"], 2)} for k, v in totals.items()}

# Strip out empty SKU expansion entries
sku_branch_stock_clean = {k: {b: round(q, 2) for b, q in v.items() if q > 0}
                           for k, v in sku_branch_stock.items()}

# Last-3m sales per SKU per branch (only SKUs that exist in stock; trim zeros)
sku_codes_in_stock = set(r["code"] for r in classified)
sku_branch_sales_3m_clean = {}
for code, bd in sku_branch_qty_last3.items():
    if code not in sku_codes_in_stock:
        continue
    pruned = {b: round(q, 2) for b, q in bd.items() if q > 0 and b in ACTIVE}
    if pruned:
        sku_branch_sales_3m_clean[code] = pruned

# Last-6m sales per SKU per branch (used by transfer bucket logic)
sku_branch_sales_6m_clean = {}
for code, bd in sku_branch_qty_last6.items():
    if code not in sku_codes_in_stock:
        continue
    pruned = {b: round(q, 2) for b, q in bd.items() if q > 0 and b in ACTIVE}
    if pruned:
        sku_branch_sales_6m_clean[code] = pruned

# Last-12m company-total sales per SKU (active branches only) — bucket logic input
sku_total_12m_clean = {}
for code, bd in sku_branch_qty_last12.items():
    if code not in sku_codes_in_stock:
        continue
    total = sum(q for b, q in bd.items() if b in ACTIVE)
    if total > 0:
        sku_total_12m_clean[code] = round(total, 2)

payload = {
    "meta": {
        "generated":    datetime.now().strftime("%Y-%m-%d %H:%M"),
        "snapshot":     ym_str(NOW),
        "source":       "202604 - SALES VS STOCK VS PO VS GRN.V2.xlsx",
        "active_branches": ACTIVE,
        "branch_names": BRANCH_NAMES,
        # 5 active store totals (used by the 5-store breakdown / dead-stock dashboard)
        "total_stock":  round(total_stock, 2),
        "problem_total": round(problem_total, 2),
        "problem_pct":  round(problem_total / total_stock * 100, 1) if total_stock else 0,
        # Company-wide totals (used by the Stock main dashboard headline number)
        "company_total_stock": round(company_total_stock, 2),
        "company_rows_total":  company_rows_total,
        "company_by_branch":   {b: round(v, 2) for b, v in company_by_branch.items()},
    },
    "totals":    totals_clean,
    "by_branch": by_branch,
    "rows":      classified,
    "sku_branch_stock": sku_branch_stock_clean,
    "sku_branch_sales_3m": sku_branch_sales_3m_clean,
    "sku_branch_sales_6m": sku_branch_sales_6m_clean,
    "sku_total_12m":       sku_total_12m_clean,
}

os.makedirs(os.path.dirname(OUT), exist_ok=True)
with open(OUT, "w", encoding="utf-8") as f:
    f.write("/* AUTO-GENERATED by tools/build-deadstock.py — DO NOT EDIT BY HAND */\n")
    f.write(f"/* Source: {os.path.basename(XLSX)} · Snapshot: {ym_str(NOW)} */\n")
    f.write("window.WP_DEADSTOCK = ")
    json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))
    f.write(";\n")

print(f"\nWrote {OUT}")
print(f"Size: {os.path.getsize(OUT):,} bytes")
print(f"Rows shipped: {len(classified)}")
